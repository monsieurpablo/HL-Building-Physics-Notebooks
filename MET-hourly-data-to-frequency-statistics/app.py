import shutil

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from PIL import Image

import utils


def read_data(WEATHER_DATA):
    # Read data from the excel file
    df = pd.read_excel(WEATHER_DATA, skiprows=10, header=0)
    # rename columns
    df.columns = ["date", "air_temp", "wet_temp", "rel_hum", "wind_speed", "wind_dir"]
    # Convert date to datetime
    df["date"] = pd.to_datetime(df["date"])
    # drop rows with NaN values on wind_dir
    df = df.dropna(subset=["wind_dir"])
    return df


def calc_wind_dir_bin(df):
    # Create wind direction bins that match the MET office format
    labels = [
        0,
        1,
        30,
        60,
        90,
        120,
        150,
        180,
        210,
        240,
        270,
        300,
        330,
    ]  # note that 0 and 1 are the same and would be corrected in the next step
    bins = pd.IntervalIndex.from_tuples(
        [
            (345, 360),
            (-0.1, 15),
            (15, 45),
            (45, 75),
            (75, 105),
            (105, 135),
            (135, 165),
            (165, 195),
            (195, 225),
            (225, 255),
            (255, 285),
            (285, 315),
            (315, 345),
        ]
    )

    # calculate frequency of each wind speed from 0 to 50 m/s every 1 m/s. after 50 m/s, goes from 50 to 999 m/s.
    wind_bin = pd.cut(df.wind_dir, bins=bins)
    wind_bin = wind_bin.cat.rename_categories(labels)
    wind_bin = wind_bin.astype("int").replace(1, 0)
    print(wind_bin.unique())
    return wind_bin


def add_margins(df):
    df.loc["ALL OBS"] = df.sum(numeric_only=True, axis=0)
    df.loc[:, "ALL OBS"] = df.sum(numeric_only=True, axis=1)
    return df


def frequency_tables(df):
    # CREATE FREQUENCY TABLES
    df_freq = pd.crosstab(df.wind_speed_round, df.wind_dir_bin).reindex(range(0, 53)).fillna(0)
    df_freq = add_margins(df_freq)
    df_freq = df_freq.astype("int")
    df_freq.loc[52] = np.nan  # create empty row at index 52

    # CREATE FREQUENCY TABLES NORMALISED
    df_freq_norm = pd.crosstab(df.wind_speed_round, df.wind_dir_bin, normalize="all").reindex(range(0, 53)).fillna(0)
    df_freq_norm = add_margins(df_freq_norm)
    df_freq_norm = (df_freq_norm * 100).round(1)
    df_freq_norm.loc[52] = np.nan  # create empty row at index 52

    return df_freq, df_freq_norm


def read_header(WEATHER_DATA):
    # read header from data file
    header = pd.read_excel(WEATHER_DATA, nrows=8, header=None).iloc[:, :2]  # read only the first two columns
    header = header.set_index(0).T

    # search text in Start and End columns after 4 digits
    start = header["Start:"].str.extract(r"\d{4}(.*)").values[0][0]
    end = header["End:"].str.extract(r"\d{4}(.*)").values[0][0]
    header["Period"] = f"{start} to {end}"

    header_str = f"""{header.Station.str.upper().values[0]}
    LAT. {header['Latitude'].values[0]} 
    LONG. {header['Longtude'].values[0]}
    ALT. {header['Altitude (m)'].values[0]}
    Period: {header['Period'].values[0]}"""

    header_lst = header_str.splitlines()
    header_lst = [x.strip() for x in header_lst]

    return header_lst


st.set_page_config(
    layout="centered",
    initial_sidebar_state="auto",  # Can be "auto", "expanded", "collapsed"
    page_title="HL | MET Office Post-processing",
    page_icon="https://i.imgur.com/BKCC4TZ.png",
)

# apply custom css
with open("style.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


image = Image.open("assets/1200px-Hoare_Lea_logo.svg_dark.png")

st.image(
    image,
    use_column_width=False,
    width=200,
)

# create an upload button in streamlit

st.title("Wind Frequency Analysis")

title_html = (
    '<p style="font-family:Roboto Slab; color:Black; font-size: 15px; text-align: left;">Pablo Arango | 2023</p>'
)
st.markdown(title_html, unsafe_allow_html=True)

st.write("---")
st.write(
    """
    **This app process the RAW year hourly date from the MET office into a monthly frequency tables used in the WDA tools.**
    """
)

st.write(
    "Upload a weather data file in the excel format provided by the MET office. A sample file can be downloaded below."
)


# Extra - Download sample file
with open("data/20-Year-Data-Northolt.xlsx", "rb") as my_file:
    st.download_button(
        label="Download Sample RAW Wind Data",
        data=my_file,
        file_name="example-RAW-MET-office-data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


WEATHER_DATA = st.file_uploader("Upload Annual Weather File", type=["xlsx"])
TEMPLATE_FILE = "data/Template_WFA.xlsx"

if WEATHER_DATA is None:
    st.warning("Please upload a file")
else:
    st.success("File uploaded successfully")

st.write("After uploading the file, click on the Process Data button to create the output file")

run = st.checkbox("Process Data")

if (run == True) and (WEATHER_DATA is not None):
    # ----------------- PROCESS DATA -----------------
    df = read_data(WEATHER_DATA)
    wind_bin = calc_wind_dir_bin(df)
    df["wind_dir_bin"] = wind_bin
    df["month"] = df.date.dt.month
    df["month_str"] = df.date.dt.strftime("%b")
    df["wind_speed_round"] = df.wind_speed.round(0)

    # ----------------- WRITE DATA TO EXCEL -----------------

    NEW_FILE = "export_wind_data_openpy.xlsx"
    # Copy the template file to a new file
    shutil.copyfile(TEMPLATE_FILE, NEW_FILE)

    # Load workbook
    wb = openpyxl.load_workbook(NEW_FILE)
    ws = wb["Sheet1"]

    # Load and format header
    header_lst = read_header(WEATHER_DATA)
    # Write header
    for row, header in enumerate(header_lst):
        ws.cell(row=row + 1, column=1).value = header

    # location list from excel template (it is not uniform)
    loc_list = [
        14,
        78,
        144,
        208,
        274,
        338,
        404,
        468,
        534,
        598,
        664,
        728,
        794,
        858,
        924,
        988,
        1054,
        1118,
        1184,
        1248,
        1314,
        1378,
        1444,
        1508,
        1574,
        1638,
    ]

    counter = 0

    # Loop through each month
    for month in range(1, 13):
        df_month = df[df.month == month]
        df_freq_month, df_freq_norm_month = frequency_tables(df_month)

        # Select initial cell and increase the row by 64 for each table for each month
        # Update workbook at specified range

        for rowid, row in enumerate(df_freq_month.values.tolist()):
            for col in range(len(row)):
                ws.cell(row=loc_list[counter] + (rowid), column=col + 2).value = row[col]

        counter += 1

        for rowid, row in enumerate(df_freq_norm_month.values.tolist()):
            for col in range(len(row)):
                ws.cell(row=loc_list[counter] + (rowid), column=col + 2).value = row[col]

        counter += 1

    # Write annual data
    df_freq, df_freq_norm = frequency_tables(df)

    # Update workbook at specified range
    for rowid, row in enumerate(df_freq.values.tolist()):
        for col in range(len(row)):
            ws.cell(row=loc_list[counter] + (rowid), column=col + 2).value = row[col]

    counter += 1

    for rowid, row in enumerate(df_freq_norm.values.tolist()):
        for col in range(len(row)):
            ws.cell(row=loc_list[counter] + (rowid), column=col + 2).value = row[col]

    counter += 1

    # Save and close workbook
    wb.save(NEW_FILE)
    wb.close()

    # ----------------- DOWNLOAD FILE -----------------

    st.success("File processed successfully")
    st.write("---")
    st.write("**Click the button below to download the processed file**")

    with open(NEW_FILE, "rb") as my_file:
        st.download_button(
            label="Download Excel File",
            data=my_file,
            file_name=NEW_FILE,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


utils.hide_streamlit_logo()
